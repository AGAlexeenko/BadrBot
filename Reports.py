import random
import time
import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime
import Variables


def refresh_reports_excel(driver, num_of_pages_to_refresh):
    workbook = load_workbook(filename='BadrBot.xlsx')
    sheet = workbook['Reports']
    last_report_date = sheet.cell(row=2+4, column=6).value  # время последнего отчёта +4 (для запаса)
    stop_parsing_reports = False  # статус остановки парсинга
    list_reports = []
    for page in range(0, num_of_pages_to_refresh):
        if stop_parsing_reports:
            break
        if page == 0:
            driver.get(Variables.server+'/report')
        else:
            driver.get(Variables.server+'/report?page=' + str(page + 1))  # прибавляем еденичку к номеру страницы
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')
        pattern = re.compile(r'sub|dat')
        items = soup.find_all('td', class_=pattern)  # ищем только элементы sub и dat
        current_date = datetime.now().strftime('%d.%m.%y')

        for item in items:

            if 'sub' in item.get('class', []):
                fight_status = item.find('img', class_=lambda x: x and 'iReport iReport' in x).get('alt')
                if fight_status == "Lost as attacker.":  # если все войска погибли, то украдено 0
                    robbed = 0
                if fight_status == "Spying was stopped successfully.":  # если разведка, то украдено 0
                    robbed = 0
                else:
                    robbed = item.find('img', class_=lambda x: x and 'reportInfo ' in x).get('alt')
                if item.find('span', class_='coordinateX'):
                    coordinateX = item.find('span', class_='coordinateX').text.strip()[1:]
                    coordinateY = item.find('span', class_='coordinateY').text.strip()[:-1]
                    coordinates0 = coordinateX + '|' + coordinateY
                    cleaned = re.sub(r'[‬‭]', '', coordinates0)
                    coordinates = cleaned.replace('−', '-')
                else:
                    a_element = item.find('div').find('a')
                    text = a_element.get_text(strip=True)
                    start_index = text.rfind(" ") + 1
                    coordinates0 = text[start_index:]
                    cleaned = re.sub(r'[‬‭]', '', coordinates0)
                    coordinates = cleaned.replace('−', '-')

                list_reports.append(coordinates)
                list_reports.append(fight_status)
                list_reports.append(robbed)
            elif 'dat' in item.get('class', []):
                dat0 = item.get_text(strip=True)
                dat = dat0.replace('today', current_date)
                list_reports.append(dat)
                if dat == last_report_date:  # если время отчёта совпадает с последним отчётов в Екселе (с учётом запаса)
                    stop_parsing_reports = True
                    break

    time.sleep(random.randint(1, 2))

    number_of_reports = int(len(list_reports) / 4)  # количество строк для вставки
    sheet.insert_rows(3, number_of_reports)  # вставить строки перед таблицей
    col = 3
    row = 3
    for report in list_reports: # записать данные в Ексель
        sheet.cell(row=row, column=col).value = report
        if col < 6:
            col += 1
        else:
            col = 3
            row += 1

    workbook.save(filename='BadrBot.xlsx')