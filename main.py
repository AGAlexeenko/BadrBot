from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium_stealth import stealth

import random
import time
import re
import csv
from openpyxl import load_workbook
from T_sounds import play_stop
from T_sounds import play_build
import tkinter as tk
from tkinter import messagebox
import sys
import T_villages
import asyncio

# from ExcelData import building_queue

service = Service(executable_path=ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
# options.add_argument("--headless") #запуск в фоновом режиме

driver = webdriver.Chrome(service=service, options=options)
stealth(driver,
        languages=["ru-RU", "ru", "en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Google Inc. (Intel)",
        renderer="ANGLE (Intel, Intel(R) HD Graphics 630 (0x0000591B) Direct3D11 vs_5_0 ps_5_0, D3D11)",
        fix_hairline=True,
        )

driver.get('https://ts5.x1.international.travian.com')

"""Авторизация"""
time.sleep(random.randint(1, 2))
login = driver.find_element(By.NAME, "name")
login.send_keys(T_auth_data.my_login)
password = driver.find_element(By.NAME, "password")
time.sleep(random.randint(1, 2))
password.send_keys(T_auth_data.my_password)
login_button = driver.find_element(By.XPATH, "//button[contains(text(),'Login')]")
time.sleep(random.randint(1, 2))
login_button.click()
time.sleep(random.randint(1, 2))

active_village = T_villages.village1
driver.get(active_village[0])
time.sleep(random.randint(1, 2))

list_of_villages = []
for v in range(1, 2):  # количество сёл
    list_of_villages.append(v)


def fHouses_Levels(village):
    """Уровни построек записываются в файл CSV"""
    fields_names_on_gids = ['noname', 'Woodcutter', 'Clay_Pit', 'Iron_Mine', 'Cropland']
    time.sleep(random.randint(1, 3))
    driver.get(village[0])

    village_name = str(
        driver.find_element(By.XPATH, f'//span[@class="coordinatesGrid" and @data-did="{village[2]}"]').get_attribute(
            'data-villagename'))

    for i in range(1, 19):
        field_slot = 'Slot' + str(i)
        full_text_field = driver.find_element(By.XPATH,
                                              f'//div[@id="resourceFieldContainer"]//a[@href="/build.php?id={i}"]').get_attribute(
            'class')
        field_gid = re.search(r'gid(.*?)buildingSlot', full_text_field).group(1).strip()
        field__name = fields_names_on_gids[int(field_gid)]
        field_level = full_text_field.split("level")[2]
        Houses_Levels[village_name + field_slot] = [field__name, field_gid, field_level]

    driver.get(village[1])
    for i in range(19, 41):
        if i < 40:
            house_gid = driver.find_element(By.XPATH,
                                            f'//div[@id="villageContent"]//div[@data-aid="{i}"]').get_attribute(
                'data-gid')
            house_slot = 'Slot' + str(i)
            if house_gid != '0':
                house_name = driver.find_element(By.XPATH,
                                                 f'//div[@id="villageContent"]//div[@data-aid="{i}"]').get_attribute(
                    'data-name')
                house_level = driver.find_element(By.XPATH,
                                                  f'//div[@id="villageContent"]//div[@data-aid="{i}"]//a[@href="/build.php?id={i}&gid={house_gid}"]').get_attribute(
                    'data-level')
                Houses_Levels[village_name + house_slot] = [house_name, house_gid, house_level]
            else:
                Houses_Levels[village_name + house_slot] = [0, 0, 0]
        else:
            try:
                house_level = driver.find_element(By.XPATH,
                                                  '//div[@id="villageContent"]//div[@data-aid="40"]//a[contains(@class, "aid40")]').get_attribute(
                    'data-level')
                Houses_Levels[village_name + 'Slot40'] = ["City Wall", 31, house_level]
            except NoSuchElementException:
                Houses_Levels[village_name + 'Slot40'] = ["City Wall", 31, 0]


Houses_Levels = {}

all_villages = (T_villages.village1,
                T_villages.village2)

result = tuple(fHouses_Levels(x) for x in all_villages)

csv_file = 'Houses_Levels.csv'
with open(csv_file, 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(['Village_name and House_Slot', 'House_name', 'House_gid', 'House_level'])  # Записываем заголовки
    for key, values in Houses_Levels.items():
        writer.writerow([key, values[0], values[1], values[2]])  # Записываем ключи и значения   

play_stop()
print("CSV с уровнями построек сохранён! Обновите файл Excel и нажмите Enter, чтобы продолжить...")
input()
print("Вы нажали Enter. Программа продолжает выполнение.")

"""Считать очередь построек из Ексель"""
dict_building_queue = {}

workbook = load_workbook(filename='TravianBot.xlsm')
sheet = workbook['Houses_Levels']

vilage_excel_first_cell = 3
vilage_excel_last_cell = 42

for village in list_of_villages:
    fields_queue0 = []
    houses_queue0 = []

    for row in range(vilage_excel_first_cell, vilage_excel_last_cell + 1):
        vil_name_and_slot = sheet.cell(row=row, column=2).value
        vil_num = sheet.cell(row=row, column=3).value
        slt_num = sheet.cell(row=row, column=4).value
        gid_num = sheet.cell(row=row, column=5).value
        house_name = sheet.cell(row=row, column=6).value
        current_level = sheet.cell(row=row, column=7).value
        times_0 = sheet.cell(row=row, column=8).value
        times1 = 0 if times_0 is None else times_0
        times = max(times1 - current_level, 0)

        while times > 0:
            current_level += 1 * 1000
            if slt_num < 19:
                fields_queue0.append(
                    f'{current_level + slt_num}https://ts5.x1.international.travian.com/dorf1.php?newdid={slt_num}&gid={gid_num}')
            else:
                houses_queue0.append(
                    f'{current_level + slt_num}https://ts5.x1.international.travian.com/dorf1.php?newdid={slt_num}&gid={gid_num}')
            times -= 1

    fields_queue0.sort()
    fields_queue1 = []
    for a in fields_queue0:
        substring = a[4:]
        fields_queue1.append(substring)
    dict_building_queue["fields_queue" + str(village)] = fields_queue1

    houses_queue0.sort()
    houses_queue1 = []

    for a in houses_queue0:
        substring = a[4:]
        houses_queue1.append(substring)

    dict_building_queue["houses_queue" + str(village)] = houses_queue1

    vilage_excel_first_cell += 40
    vilage_excel_last_cell += 40


def ordered_buildings():
    # """Количество построек в очереди"""
    try:
        building_list = driver.find_element(By.XPATH, '//div[@class="buildingList"]').text
        words_in_building_list = building_list.split()
        count_buildings_in_list = sum(1 for word in words_in_building_list if word == "Level")
        timer_building = driver.find_element(By.XPATH,
                                             '//div[@class="buildDuration"]//span[@class="timer"]').get_attribute(
            'value')
        # print("Построек в очереди"+str(count_buildings_in_list))
        return count_buildings_in_list, timer_building
    except NoSuchElementException:
        return 0, 0


async def building(village, fields_queue, houses_queue):
    '''
    DOCSTRING: Заказать постройку
    INPUT: Название постройки
    OUTPUT: Сообщение об окончании постройки
    '''
    time.sleep(random.randint(1, 3))
    driver.get(village[0])
    time.sleep(random.randint(1, 3))
    house_index = 0

    for field_adress in fields_queue:

        count_buildings_in_list, timer_building = ordered_buildings()
        print(count_buildings_in_list)

        if count_buildings_in_list == 3:
            print("Ждём " + timer_building + " секунд")
            await asyncio.sleep(int(timer_building) + random.randint(1, 5))
            driver.get(village[0])
            time.sleep(random.randint(1, 3))

        driver.get(field_adress)
        try:
            time.sleep(random.randint(1, 3))
            build_button = driver.find_element(By.XPATH, '//button[@class="textButtonV1 green build"]')
            building_name = driver.find_element(By.XPATH, '//h1[@class="titleInHeader"]').text
            time.sleep(random.randint(1, 3))
            build_button.click()
            play_build()
            print("заказано поле " + building_name + ' в деревне ' + str(village[3]))

        except NoSuchElementException:
            time.sleep(random.randint(1, 3))
            driver.get(houses_queue[house_index])
            try:
                build_button = driver.find_element(By.XPATH, '//button[@class="textButtonV1 green build"]')
                building_name = driver.find_element(By.XPATH, '//h1[@class="titleInHeader"]').text
                time.sleep(random.randint(1, 3))
                build_button.click()
                play_build()
                house_index += 1
                print("заказан дом " + building_name + ' в деревне ' + str(village[3]))
            except NoSuchElementException:
                time.sleep(random.randint(1, 3))
                # потом доработать  это ПАС ошибки "не хватает ресурсов"   //div[@class="errorMessage"]
                if True:
                    driver.get('https://ts5.x1.international.travian.com/hero/inventory')
                    time.sleep(random.randint(1, 3))
                    Inventar_wood_button = driver.find_element(By.XPATH,
                                                               '//div[@data-slot="inventory" and @data-placeid="1"]')
                    Inventar_wood_button.click()
                    time.sleep(random.randint(1, 3))
                    try:
                        Transfer_button = driver.find_element(By.XPATH,
                                                              '//button[@class="textButtonV2 buttonFramed rectangle withText green"]')
                        Transfer_button.click()
                    except NoSuchElementException:
                        print('Заполнено')
                    time.sleep(random.randint(1, 3))
                    Inventar_clay_button = driver.find_element(By.XPATH,
                                                               '//div[@data-slot="inventory" and @data-placeid="2"]')
                    Inventar_clay_button.click()
                    time.sleep(random.randint(1, 3))
                    try:
                        Transfer_button = driver.find_element(By.XPATH,
                                                              '//button[@class="textButtonV2 buttonFramed rectangle withText green"]')
                        Transfer_button.click()
                    except NoSuchElementException:
                        print('Заполнено')

                    time.sleep(random.randint(1, 3))
                    Inventar_iron_button = driver.find_element(By.XPATH,
                                                               '//div[@data-slot="inventory" and @data-placeid="3"]')
                    Inventar_iron_button.click()
                    time.sleep(random.randint(1, 3))
                    try:
                        Transfer_button = driver.find_element(By.XPATH,
                                                              '//button[@class="textButtonV2 buttonFramed rectangle withText green"]')
                        Transfer_button.click()
                    except NoSuchElementException:
                        print('Заполнено')
                    time.sleep(random.randint(1, 3))
                    Inventar_crop_button = driver.find_element(By.XPATH,
                                                               '//div[@data-slot="inventory" and @data-placeid="4"]')
                    Inventar_crop_button.click()
                    time.sleep(random.randint(1, 3))
                    try:
                        Transfer_button = driver.find_element(By.XPATH,
                                                              '//button[@class="textButtonV2 buttonFramed rectangle withText green"]')
                        Transfer_button.click()
                    except NoSuchElementException:
                        print('Заполнено')
                    driver.get(field_adress)
                    try:
                        time.sleep(random.randint(1, 3))
                        build_button = driver.find_element(By.XPATH, '//button[@class="textButtonV1 green build"]')
                        building_name = driver.find_element(By.XPATH, '//h1[@class="titleInHeader"]').text
                        time.sleep(random.randint(1, 3))
                        build_button.click()
                        play_build()
                        print("заказано поле " + building_name + ' в деревне ' + str(village[3]))

                    except NoSuchElementException:
                        time.sleep(random.randint(1, 3))
                        driver.get(houses_queue[house_index])
                        build_button = driver.find_element(By.XPATH, '//button[@class="textButtonV1 green build"]')
                        building_name = driver.find_element(By.XPATH, '//h1[@class="titleInHeader"]').text
                        # building_level = driver.find_element(By.XPATH, '//h1[@class="titleInHeader"]//span[@class="level"]').text[:8]
                        time.sleep(random.randint(1, 3))
                        build_button.click()
                        play_build()
                        house_index += 1
                        print("заказан дом " + building_name + ' в деревне ' + str(village[3]))

                else:
                    play_stop()

                    def continue_clicked():
                        messagebox.showinfo("Продолжение", "Вы продолжили работу")
                        # root.quit()
                        root.destroy()

                    def stop_clicked():
                        messagebox.showwarning("Остановка", "Программа остановлена")
                        root.quit()
                        sys.exit()

                    root = tk.Tk()
                    root.title("Запрос к пользователю")

                    # Создаем метку с инструкцией
                    label = tk.Label(root, text="не могу построить - не хватает ресов. И нажмите Enter")
                    label.pack(pady=10)

                    # Создаем кнопку для продолжения
                    continue_button = tk.Button(root, text="Продолжить", command=continue_clicked)
                    continue_button.pack(pady=5)

                    # Создаем кнопку для остановки
                    stop_button = tk.Button(root, text="Остановить программу", command=stop_clicked)
                    stop_button.pack(pady=5)

                    def lift_message_window():
                        root.after(100, root.lift)

                        # Вызываем функцию поднятия окна сообщения через 100 миллисекунд

                    root.after(100, lift_message_window)

                    # Запускаем главный цикл обработки событий
                    root.mainloop()
                    # print(f"не могу построить{key}: не хватает ресов или закажите здание внутри. И нажмите Enter")
                    # input()
                    # print("Вы нажали Enter. Программа продолжает выполнение.")
                    driver.get(field_adress)
                    try:
                        time.sleep(random.randint(1, 3))
                        build_button = driver.find_element(By.XPATH, '//button[@class="textButtonV1 green build"]')
                        build_button.click()
                        print("заказано")

                    except NoSuchElementException:
                        time.sleep(random.randint(1, 3))
                        driver.get(houses_queue[house_index])
                        build_button = driver.find_element(By.XPATH, '//button[@class="textButtonV1 green build"]')
                        build_button.click()
                        house_index += 1
                        print("Дом заказан")


async def main():
    await asyncio.gather(
        building(T_villages.village1, dict_building_queue['fields_queue1'], dict_building_queue['houses_queue1']),
        building(T_villages.village1, dict_building_queue['fields_queue1'], dict_building_queue['houses_queue1'])
        )


asyncio.run(main())

# building(fields_queue, houses_queue)
# building(fields_queue1, houses_queue1)
print("Успешно окончено!")