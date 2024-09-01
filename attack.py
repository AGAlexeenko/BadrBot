from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium_stealth import stealth
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import random
import time
import tkinter as tk
from tkinter import ttk
import re
from PIL import Image, ImageTk
import threading
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import OrderedDict
import Oasises
import Variables
from RallyPoint import parsing_rally_point
from Reports import refresh_reports_excel

if Variables.race == 'Romans':
    troops = ['', 'Legionnaire','Praetorian','Imperian','Equites Legati', 'Equites Imperatoris','Equites Caesaris','empty','empty','empty','empty','Hero']
else:
    troops = ['', 'Mercenary', 'Bowman', 'Spotter', 'Steppe Rider', 'Marksman','Marauder','empty','empty','empty','empty','Hero']

active_village = []
building_queue_dic = {}
dic_all_villages = {}

def open_browser():
    global driver
    global wait
    global active_village

    service = Service(executable_path='venv/chromedriver-win64/chromedriver.exe')
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    # options.add_argument("--headless") #запуск в фоновом режиме
    driver = webdriver.Chrome(service=service, options=options)

    stealth(driver,
            languages=["ru-RU", "ru", "en-US", "en"],
            vendor="Google Inc.",
            platform="tab132",
            webgl_vendor="Google Inc. (Intel)",
            renderer="ANGLE (Intel, Intel(R) HD Graphics 630 (0x0000591B) Direct3D11 vs_5_0 ps_5_0, D3D11)",
            fix_hairline=True,
            )
    wait = WebDriverWait(driver, 15, poll_frequency=1)
    driver.get(Variables.server)

    """Авторизация"""
    time.sleep(random.uniform(0.5, 1))
    driver.find_element("name", "name").send_keys(Variables.my_login)
    time.sleep(random.uniform(0.5, 1))
    driver.find_element("name", "password").send_keys(Variables.my_password)
    time.sleep(random.uniform(0.5, 1))
    driver.find_element("xpath",
                        "//button[@class='textButtonV2 buttonFramed rectangle withText withLoadingIndicator green']").click()

    """список деревень"""
    # wait.until(EC.visibility_of_element_located(("xpath", '//div[@class="boxTitle"]')))  # ждёт пока не загрузится
    time.sleep(random.randint(1, 2))
    all_villages_elements = driver.find_elements("xpath", '//span[@class="coordinatesGrid"]')

    for element in all_villages_elements:
        village_id = element.get_attribute('data-did')
        village_name = element.get_attribute('data-villagename')
        village_x = element.get_attribute('data-x')
        village_y = element.get_attribute('data-y')
        fields_url = Variables.server + '/dorf1.php?newdid=' + village_id + '&'  # адресс url снаружи села
        houses_url = Variables.server + '/dorf2.php?newdid=' + village_id + '&'  # адресс url внутри села
        dic_all_villages[village_name] = [village_id, village_x, village_y, fields_url, houses_url]
    time.sleep(random.uniform(0.5, 1))
    active_village = dic_all_villages["2"]  # активное село
    driver.get(active_village[3])


def on_button_refresh_reports_excel_click():
    num_of_pages_to_refresh = int(num_of_pages_refresh_reports_excel.get())
    refresh_reports_excel(driver, num_of_pages_to_refresh)
    parsing_rally_point(driver)


def send_troops_new_window(type_of_troops1, type_of_troops2="0", type_of_troops3="0"):  # Отправить войска в новом окне

    url = driver.current_url
    matches = re.findall(r'x=(-?\d+)&y=(-?\d+)', url)
    key_for_x = matches[0][0]
    key_for_y = matches[0][1]
    elements = driver.find_elements("xpath", '//a[@class="a arrow"]')  # центрировать, атаковать, симулятор
    link = elements[1].get_attribute('href')  # отбираем 2-й элемент - атаковать
    driver.switch_to.new_window('tab')  # открываем новую вкладку
    driver.get(link)

    number_of_troops = get_number_of_troops(type_of_troops1)
    if type_of_troops2 != "0":
        number_of_troops2 = get_number_of_troops(type_of_troops2)
    else:
        number_of_troops2 = 0
    if type_of_troops3 != "0":
        number_of_troops3 = get_number_of_troops(type_of_troops3)
    else:
        number_of_troops3 = 0

    name_of_troops = troops[int(type_of_troops1)]
    name_of_troops2 = troops[int(type_of_troops2)]
    name_of_troops3 = troops[int(type_of_troops3)]

    def close_new_window_and_karte():  # закрыть мини окно карты и вернуться в основное окно
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element("xpath", '//div[@class="dialogCancelButton iconButton small cancel"]').click()

    if number_of_troops == 1 and type_of_troops1 == '1':
        update_label_message(lable_message,
                             'один воин умрёт!')
        close_new_window_and_karte()

    try:  # проверяет, сколько войск дома 1-го типа войск
        troops_at_home_text = driver.find_element("xpath", f'//a[contains(@onclick, "troop[t{type_of_troops1}]")]').text
        if type_of_troops2 != "0":
            troops2_at_home_text = driver.find_element("xpath", f'//a[contains(@onclick, "troop[t{type_of_troops2}]")]').text
        if type_of_troops3 != "0":
            troops3_at_home_text = driver.find_element("xpath", f'//a[contains(@onclick, "troop[t{type_of_troops3}]")]').text
    except NoSuchElementException:  # если войск нет вообще
        print(f'в пункте нет войск для {key_for_x}|{key_for_y}')
        close_new_window_and_karte()
        return

    troops_at_home_num = int(''.join(filter(str.isdigit, troops_at_home_text)))
    if type_of_troops2 != "0":
        troops2_at_home_num = int(''.join(filter(str.isdigit, troops2_at_home_text)))
    else:
        troops2_at_home_num = 1
    if type_of_troops3 != "0":
        troops3_at_home_num = int(''.join(filter(str.isdigit, troops3_at_home_text)))
    else:
        troops3_at_home_num = 1

    if number_of_troops > troops_at_home_num or number_of_troops2 > troops2_at_home_num or number_of_troops3 > troops3_at_home_num:  # если в отправке войск больше, чем есть дома
        update_label_message(lable_message,
                             f'не хватает войск для {key_for_x}|{key_for_y}')
        close_new_window_and_karte()
    else:
        driver.find_element("name", f"troop[t{type_of_troops1}]").send_keys(str(number_of_troops))
        if type_of_troops2 != "0":
            driver.find_element("name", f"troop[t{type_of_troops2}]").send_keys(str(number_of_troops2))
        if type_of_troops3 != "0":
            driver.find_element("name", f"troop[t{type_of_troops3}]").send_keys(str(number_of_troops3))
        driver.find_element("name", "ok").click()  # send button
        confirm_locator = ("xpath", '//button[@class="textButtonV1 green rallyPointConfirm"]')
        wait.until(EC.visibility_of_element_located(confirm_locator)).click()
        if type_of_troops2 != "0" and type_of_troops3 != "0":
            print(f'{number_of_troops} {name_of_troops} и {number_of_troops2} {name_of_troops2} и {number_of_troops3} {name_of_troops3}  отправлены в рейд в {key_for_x}|{key_for_y}')
        elif type_of_troops2 != "0":
            print(f'{number_of_troops} {name_of_troops} и {number_of_troops2} {name_of_troops2} отправлены в рейд в {key_for_x}|{key_for_y}')
        else:
            print(f'{number_of_troops} {name_of_troops} отправлены в рейд в {key_for_x}|{key_for_y}')
        time.sleep(random.uniform(0.5, 1))
        close_new_window_and_karte()


def btn_send_troops_timer(type_of_troops):
    url = driver.current_url
    matches = re.findall(r'x=(-?\d+)&y=(-?\d+)', url)
    key_for_x = matches[0][0]
    key_for_y = matches[0][1]
    troops_sender_timer = get_troops_sender_timer(type_of_troops)
    number_of_troops = get_number_of_troops(type_of_troops)
    driver.find_element("xpath", '//div[@class="dialogCancelButton iconButton small cancel"]').click()
    print(f'ожидают атаки {number_of_troops} {type_of_troops} на {key_for_x}|{key_for_y}')
    threading.Timer(troops_sender_timer, send_troops_timer, args=(type_of_troops, key_for_x, key_for_y, number_of_troops)).start()


def on_btn_build_after_or_in_time():
    url = driver.current_url
    village_name = driver.find_element("xpath", '//input[@name="villageName"]').get_attribute("value")
    field_and_level_name = driver.find_element("xpath", '//h1[@class="titleInHeader"]').text
    field_name = field_and_level_name.split('Level')[0].strip()
    next_possible_level = driver.find_element("xpath", '//tr[@class="nextPossible"]').text.split('level')[-1].split(':')[0].strip()
    after_timer = int(Entry_build_after_time.get())  # секунды, через которые начнётся стройка
    in_time = Entry_build_in_time.get()  # время, в которое начнётся стройка
    hours, minutes = map(int, in_time.split(':'))
    if not (0 <= hours < 24 and 0 <= minutes < 60):
        print("Ошибка: Неправильный формат времени")
        return
    now = datetime.now()
    target_time = now.replace(hour=hours, minute=minutes, second=0, microsecond=0)
    time_difference = target_time - now
    seconds_difference = time_difference.total_seconds()
    if in_time == '00:00':
        final_seconds = after_timer
        final_time = now + timedelta(seconds=final_seconds)
        print(f'построим {field_name} {next_possible_level} в {final_time.strftime("%H:%M:%S")}')
    else:
        final_seconds = seconds_difference
        final_time = now + timedelta(seconds=final_seconds)
        print(f'построим {field_name} {next_possible_level} в {final_time.strftime("%H:%M:%S")}')
    unic_time = final_time.strftime("%H:%M:%S")
    building_queue_dic[unic_time] = [village_name, field_name, next_possible_level, final_time.strftime("%H:%M:%S")]
    sorted_building_queue_dic = OrderedDict(sorted(building_queue_dic.items()))
    display_building_queue_listbox(sorted_building_queue_dic)
    threading.Timer(final_seconds, build_field, args=(url,village_name,unic_time)).start()


def build_field(url, vil_name, unic_time):  # построить здание
    village_name = driver.find_element("xpath", '//input[@name="villageName"]').get_attribute("value")
    if village_name != vil_name: # проверить, в том ли селе находимся
        active_village = dic_all_villages[vil_name]  # получить список элементов нужного села
        driver.get(active_village[3])  # сменить село
        time.sleep(random.uniform(0.5, 2))
    i = 0
    random_start = 5  # время ожидания +10 секунд, и каждую попытку умножается на 3
    while i < 3:  # 3 попытки
        try:
            driver.get(url)
            time.sleep(random.uniform(0.5, 2))
            element = driver.find_element("xpath", '//button[@class="textButtonV1 green build"]')
            element.click()
            building_queue_dic[unic_time].append(" - заказано")
            sorted_building_queue_dic = OrderedDict(sorted(building_queue_dic.items()))
            display_building_queue_listbox(sorted_building_queue_dic)
            break
        except NoSuchElementException:  # если кнопка не найдена
            building_queue_dic[unic_time].append(" - не смог! Пытаюсь ещё!")
            sorted_building_queue_dic = OrderedDict(sorted(building_queue_dic.items()))
            display_building_queue_listbox(sorted_building_queue_dic)
            time.sleep(random.randint(random_start, random_start+10))
            i += 1
            if i == 3:
                building_queue_dic[unic_time].append(" - не заказано!")
                sorted_building_queue_dic = OrderedDict(sorted(building_queue_dic.items()))
                display_building_queue_listbox(sorted_building_queue_dic)
            random_start *= 3


def send_troops_timer(type_of_troops, key_for_x, key_for_y, number_of_troops):
    times_to_repeat = int(btn_farm_wait_times_to_repeat.get())  # количество повторений, если нет или тсутствуют войска

    if type_of_troops == "1":
        repeat = 0

        while repeat < times_to_repeat:
            try:
                driver.get(Variables.raid)
                troops_at_home_text = driver.find_element("xpath", '//a[contains(@onclick, "troop[t1]")]').text
                break
            except NoSuchElementException:
                if times_to_repeat > 4:
                    seconds_to_wait = random.randint(180, 600)
                    print(
                        f'в пункте нет войск {type_of_troops} {number_of_troops} ...ждём {seconds_to_wait}')
                    time.sleep(seconds_to_wait)
                elif times_to_repeat > 1:
                    seconds_to_wait = random.randint(5, 60)
                    print(
                        f'в пункте нет войск {type_of_troops} {number_of_troops} ...ждём {seconds_to_wait}')
                    time.sleep(seconds_to_wait)
                repeat += 1

        troops_at_home_num = int(''.join(filter(str.isdigit, troops_at_home_text)))  # оставить только цифры
        repeat = 0
        while number_of_troops > troops_at_home_num and repeat < times_to_repeat:
            if times_to_repeat == 1:
                time.sleep(random.uniform(0.5, 1))
                continue

            if times_to_repeat > 4:
                seconds_to_wait = random.randint(180, 600)
                print(
                    f'в пункте не хватает войск {type_of_troops} {number_of_troops} ...ждём {seconds_to_wait}')
                time.sleep(seconds_to_wait)
            elif times_to_repeat > 1:
                seconds_to_wait = random.randint(5, 60)
                print(
                    f'в пункте не хватает войск {type_of_troops} {number_of_troops} ...ждём {seconds_to_wait}')
                time.sleep(seconds_to_wait)
            driver.get(Variables.raid)
            time.sleep(random.randint(1, 2))
            troops_at_home_text = driver.find_element("xpath", '//a[contains(@onclick, "troop[t1]")]').text
            try:
                troops_at_home_num = int(''.join(filter(str.isdigit, troops_at_home_text)))
            except ValueError as ex:
                cleaned_string = ''.join(filter(str.isdigit, troops_at_home_num))
                troops_at_home_num = int(cleaned_string)
            repeat += 1

        driver.find_element("name", "troop[t1]").send_keys(str(number_of_troops))

    driver.find_element("id", "xCoordInput").send_keys(key_for_x)
    driver.find_element("id", "yCoordInput").send_keys(key_for_y)
    driver.find_element("name", "ok").click()

    confirm_locator = ("xpath", '//button[@class="textButtonV1 green rallyPointConfirm"]')
    wait.until(EC.visibility_of_element_located(confirm_locator)).click()
    print(f'{number_of_troops} {type_of_troops} отправлены в рейд')
    time.sleep(random.uniform(0.5, 1))


def get_number_of_troops(type_of_troops):  # получить тип войск
    if type_of_troops == "1":
        return int(ent_troops1.get())
    if type_of_troops == "4":
        return int(ent_troops4.get())
    if type_of_troops == "5":
        return int(troops5.get())
    if type_of_troops == "11":
        return 1


def get_troops_sender_timer(type_of_troops):
    if type_of_troops == "1":
        return int(troops1_sender_timer.get())
    if type_of_troops == "5":
    #     return int(troops5_sender_timer.get())
    # if type_of_troops == "11":
        return int(troops_hero_sender_timer.get())


def add_to_farmlist(type_of_troops1, type_of_troops2=0, type_of_troops3=0):
    url = driver.current_url
    matches = re.findall(r'x=(-?\d+)&y=(-?\d+)', url)
    key_for_x = matches[0][0]
    key_for_y = matches[0][1]
    cordinates = key_for_x + '|' + key_for_y

    farm_title0 = driver.find_element("xpath", '//h1[@class ="titleInHeader"]').text
    farm_title = farm_title0.split('(')[0]
    elements = driver.find_elements("xpath", '//a[@class="a arrow"]')
    element = elements[1].get_attribute('href')
    Map_id = element.split('MapId=')[1]

    x_home = driver.find_element("xpath", f'//span[@class="coordinatesGrid" and @data-did="{str(active_village[0])}"]').get_attribute(
        'data-x')
    y_home = driver.find_element("xpath", f'//span[@class="coordinatesGrid" and @data-did="{str(active_village[0])}"]').get_attribute(
        'data-y')

    distance = round(((int(key_for_x) - int(x_home)) ** 2 + (int(key_for_y) - int(y_home)) ** 2) ** 0.5, 2)
    if type_of_troops1 != "0":
        number_of_troops1 = get_number_of_troops(type_of_troops1)
    if type_of_troops2 != "0":
        number_of_troops2 = get_number_of_troops(type_of_troops2)
    if type_of_troops3 != "0":
        number_of_troops3 = get_number_of_troops(type_of_troops3)

    workbook = load_workbook(filename='BadrBot.xlsx')
    sheet = workbook['Farmlist']
    last_row = sheet.max_row

    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row=row, column=3).value is not None:
            last_row = row
            break
    sheet['B' + str(last_row + 1)] = Map_id
    sheet['C' + str(last_row + 1)] = farm_title
    sheet['D' + str(last_row + 1)] = cordinates
    sheet['E' + str(last_row + 1)] = key_for_x
    sheet['F' + str(last_row + 1)] = key_for_y
    if type_of_troops1 != "0":
        sheet['H' + str(last_row + 1)] = troops[int(type_of_troops1)]
        sheet['I' + str(last_row + 1)] = number_of_troops1
    if type_of_troops2 != "0":
        sheet['G' + str(last_row + 1)] = troops[int(type_of_troops2)]
        sheet['K' + str(last_row + 1)] = number_of_troops2
    if type_of_troops3 != "0":
        sheet['L' + str(last_row + 1)] = troops[int(type_of_troops3)]
        sheet['M' + str(last_row + 1)] = number_of_troops3
    sheet['N' + str(last_row + 1)] = distance
    update_label_message(lable_add_to_farmlist_status, f"{cordinates} добавлена в фармлист")
    driver.find_element("xpath", '//div[@class="dialogCancelButton iconButton small cancel"]').click()
    workbook.save(filename='BadrBot.xlsx')


def start_farmlist(excel_sheet):
    """
    DOCSTRING: При отсутствии или нехватке войск, если указано кол-во повторений:
    "0" - то отправляться следующая очередь
    до "5" - то время ожидания от 5 до 60 секунд
    5 и более - то время ожидания 180-600 секунд
    """
    times_to_repeat = int(btn_farm_wait_times_to_repeat.get())  # количество повторений, если нет или осутствуют войска
    farm_circles = int(ent_farm_circles.get())
    workbook = load_workbook(filename='BadrBot.xlsx')
    sheet = workbook[excel_sheet]

    last_row = sheet.max_row
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row=row, column=3).value is not None:
            last_row = row
            break

    farmlist_of_lists = []
    type_of_troops1_xl = 0
    num_of_troops1_xl = 0
    type_of_troops2_xl = 0
    num_of_troops2_xl = 0
    type_of_troops3_xl = 0
    num_of_troops3_xl = 0

    for row in range(4, last_row + 1):  # считать все данные с Эксель листа
        priority = sheet.cell(row=row, column=7).value
        if priority is None or priority == 0:  # если приоритет не задан, то пропустить строку
            continue
        x_xl = sheet.cell(row=row, column=5).value
        y_xl = sheet.cell(row=row, column=6).value

        type_of_troops1_xl = sheet.cell(row=row, column=8).value
        num_of_troops1_xl = sheet.cell(row=row, column=9).value
        if type_of_troops1_xl is None or num_of_troops1_xl is None:
            type_of_troops1_xl = 0
            num_of_troops1_xl = 0
        else:
            type_of_troops1_xl = troops.index(type_of_troops1_xl)

        type_of_troops2_xl = sheet.cell(row=row, column=10).value
        num_of_troops2_xl = sheet.cell(row=row, column=11).value
        if type_of_troops2_xl is None or num_of_troops2_xl is None:
            type_of_troops2_xl = 0
            num_of_troops2_xl = 0
        else:
            type_of_troops2_xl = troops.index(type_of_troops2_xl)

        type_of_troops3_xl = sheet.cell(row=row, column=12).value
        num_of_troops3_xl = sheet.cell(row=row, column=13).value
        if type_of_troops3_xl is None or num_of_troops3_xl is None:
            type_of_troops3_xl = 0
            num_of_troops3_xl = 0
        else:
            type_of_troops3_xl = troops.index(type_of_troops3_xl)

        farmlist_of_lists.append([priority, x_xl, y_xl, type_of_troops1_xl, num_of_troops1_xl, type_of_troops2_xl, num_of_troops2_xl, type_of_troops3_xl, num_of_troops3_xl])

    farmlist_of_lists_one_circle = sorted(farmlist_of_lists, key=lambda x: x[0],
                               reverse=True)  # отсортировать по убыванию приоритета
    farmlist_of_lists = farmlist_of_lists_one_circle*farm_circles

    quantity_of_farms = len(farmlist_of_lists)
    stop_farm_counter = 0  # счётчик для остановки фарм машины
    send_count = 0  # счётчик отправленных атак

    for farm in farmlist_of_lists:
        if stop_farm_counter > 200:  # если более 3х раз нехватило войск, то фарммашина останавливается
            print(f"Фармашина отработала не полностью из-за нехватки войск. Остановлена на {farm}")
            break
        x_farm = farm[1]
        y_farm = farm[2]
        type_of_troops1_farm = farm[3]
        name_of_troops1 = troops[int(type_of_troops1_farm)]
        num_of_troops1_farm = int(farm[4])
        type_of_troops2_farm = farm[5]
        name_of_troops2 = troops[int(type_of_troops2_farm)]
        num_of_troops2_farm = int(farm[6])
        type_of_troops3_farm = farm[7]
        name_of_troops3 = troops[int(type_of_troops3_farm)]
        num_of_troops3_farm = int(farm[8])

        if num_of_troops1_farm == 1 and type_of_troops1_farm == '1':  # проверка, чтобы не отправить 1 слабого воина
            print('один воин умрёт')
            continue

        repeat = 0
        while repeat <= times_to_repeat:  # не более указанного количества повторений
            try:
                driver.get(Variables.raid)
                time.sleep(random.uniform(0.5, 1.5))
                if type_of_troops1_farm != 0 and num_of_troops1_farm != 0:
                    troops1_at_home_text = driver.find_element("xpath", f'//a[contains(@onclick, "troop[t{type_of_troops1_farm}]")]').text
                if type_of_troops2_farm != 0 and num_of_troops2_farm !=0:
                    troops2_at_home_text = driver.find_element("xpath",
                                                           f'//a[contains(@onclick, "troop[t{type_of_troops2_farm}]")]').text
                if type_of_troops3_farm != 0 and num_of_troops3_farm !=0:
                    troops3_at_home_text = driver.find_element("xpath",
                                                           f'//a[contains(@onclick, "troop[t{type_of_troops3_farm}]")]').text
                break  # если войск хватает, то пропускаем ожидание
            except NoSuchElementException:
                if times_to_repeat > 4:  # если указано повторений > 4, то ждать будет до 10 минут
                    seconds_to_wait = random.randint(180, 360)
                    print(
                        f'в пункте нет войск для {x_farm}|{y_farm} ...ждём {seconds_to_wait}')
                    time.sleep(seconds_to_wait)
                elif times_to_repeat > 0:  # если указано повторений > 0, то ждать будет до минуты
                    seconds_to_wait = random.randint(180, 360)
                    print(
                        f'в пункте нет войск для {x_farm}|{y_farm} ...ждём {seconds_to_wait}')
                    time.sleep(seconds_to_wait)
                repeat += 1  # если указано 0 повторений, то повторений не будет

        if repeat > times_to_repeat and go_to_next_farm.get():
            print(f'цель {x_farm}|{y_farm} пропущена из-за отсутствия войск')
            continue

        if repeat > times_to_repeat and not go_to_next_farm.get():  # если повторений было больше, чем указано, значит атака не отправлена. И если не задано "переходить к следующей цели", то увеличиваем счётчик остановки фарммашины
            time.sleep(random.uniform(0.5, 1))
            stop_farm_counter += 1
            continue
        if type_of_troops1_farm ==0 or num_of_troops1_farm ==0:
            troops1_at_home_num = 1
        else:
            troops1_at_home_num = int(''.join(filter(str.isdigit, troops1_at_home_text)))  # оставить только цифры

        if type_of_troops2_farm ==0 or num_of_troops2_farm ==0:
            troops2_at_home_num = 1
        else:
            troops2_at_home_num = int(''.join(filter(str.isdigit, troops2_at_home_text)))  # оставить только цифры

        if type_of_troops3_farm ==0 or num_of_troops3_farm ==0:
            troops3_at_home_num = 1
        else:
            troops3_at_home_num = int(''.join(filter(str.isdigit, troops3_at_home_text)))  # оставить только цифры

        repeat = 0

        if times_to_repeat == 0 and num_of_troops1_farm > troops1_at_home_num and num_of_troops2_farm > troops2_at_home_num and num_of_troops3_farm > troops3_at_home_num and repeat <= times_to_repeat and not go_to_next_farm.get():
            print(f'в пункте не хватает войск ...отмена {x_farm}|{y_farm}')
            time.sleep(random.uniform(0.5, 1))
            stop_farm_counter += 1
            continue

        while num_of_troops1_farm > troops1_at_home_num and num_of_troops2_farm > troops2_at_home_num and num_of_troops3_farm > troops3_at_home_num and repeat <= times_to_repeat and go_to_next_farm.get():

            if times_to_repeat > 4:
                seconds_to_wait = random.randint(180, 3600)
                print(
                    f'в пункте не хватает войск для {x_farm}|{y_farm}...ждём {seconds_to_wait}')
                time.sleep(seconds_to_wait)
            elif times_to_repeat > 0:
                seconds_to_wait = random.randint(180, 360)
                print(
                    f'в пункте не хватает войск для {x_farm}|{y_farm} ...ждём {seconds_to_wait}')
                time.sleep(seconds_to_wait)
            driver.get(Variables.raid)
            time.sleep(random.randint(1, 2))

            if type_of_troops1_farm !=0 and num_of_troops1_farm !=0:
                troops1_at_home_text = driver.find_element("xpath",
                                                       f'//a[contains(@onclick, "troop[t{type_of_troops1_farm}]")]').text
            if type_of_troops2_farm !=0 and num_of_troops2_farm !=0:
                troops2_at_home_text = driver.find_element("xpath",
                                                       f'//a[contains(@onclick, "troop[t{type_of_troops2_farm}]")]').text
            if type_of_troops3_farm !=0 and num_of_troops3_farm !=0:
                troops3_at_home_text = driver.find_element("xpath",
                                                       f'//a[contains(@onclick, "troop[t{type_of_troops3_farm}]")]').text
            try:
                troops1_at_home_num = int(''.join(filter(str.isdigit, troops1_at_home_text)))
                troops2_at_home_num = int(''.join(filter(str.isdigit, troops2_at_home_text)))
                troops3_at_home_num = int(''.join(filter(str.isdigit, troops3_at_home_text)))
            except ValueError as ex:
                cleaned_string = ''.join(filter(str.isdigit, troops1_at_home_num))
                troops1_at_home_num = int(cleaned_string)
                cleaned_string = ''.join(filter(str.isdigit, troops2_at_home_num))
                troops2_at_home_num = int(cleaned_string)
                cleaned_string = ''.join(filter(str.isdigit, troops3_at_home_num))
                troops3_at_home_num = int(cleaned_string)
            repeat += 1

        if repeat > times_to_repeat and go_to_next_farm.get():
            print(f'цель {x_farm}|{y_farm} пропущена из-за нехватки войск')
            continue

        if repeat > times_to_repeat and not go_to_next_farm.get():  # если повторений было больше, чем указано, значит атака не отправлена. И если не задано "переходить к следующей цели", то увеличиваем счётчик остановки фарммашины
            time.sleep(random.uniform(0.5, 1))
            print(f'цель {x_farm}|{y_farm} пропущена из-за нехватки войск')
            stop_farm_counter += 1
            continue
        if type_of_troops1_farm != 0 and num_of_troops1_farm != 0:
            driver.find_element("name", f"troop[t{type_of_troops1_farm}]").send_keys(str(num_of_troops1_farm))
        if type_of_troops2_farm != 0 and num_of_troops2_farm != 0:
            driver.find_element("name", f"troop[t{type_of_troops2_farm}]").send_keys(str(num_of_troops2_farm))
        if type_of_troops3_farm != 0 and num_of_troops3_farm != 0:
            driver.find_element("name", f"troop[t{type_of_troops3_farm}]").send_keys(str(num_of_troops3_farm))

        driver.find_element("id", "xCoordInput").send_keys(x_farm)
        driver.find_element("id", "yCoordInput").send_keys(y_farm)
        driver.find_element("name", "ok").click()

        confirm_locator = ("xpath", '//button[@class="textButtonV1 green rallyPointConfirm"]')
        wait = WebDriverWait(driver, 15, poll_frequency=1)
        wait.until(EC.visibility_of_element_located(confirm_locator)).click()
        print(f'Войска отправлены в рейд в {x_farm}|{y_farm}')
        send_count += 1
        time.sleep(random.uniform(0.5, 5))
        update_label_message(lable_message, f"...отправлено {send_count} из {quantity_of_farms} целей")
    update_label_message(lable_message, f"Всё! Отправлено {send_count} из {quantity_of_farms} целей")


def update_label_message(lable, new_text):
    lable.config(text=new_text)


def display_building_queue_listbox(queue_dic):

    building_queue_listbox.delete(0, tk.END)

    for value in queue_dic.values():
        building_queue_listbox.insert(tk.END, value)

win = tk.Tk()
win.title('BadrBot_V1')
win.geometry("320x380+1390+40")
win.attributes('-topmost', True)

notebook = ttk.Notebook(win)
notebook.pack(fill='both', expand=True)

tab1 = ttk.Frame(notebook)
notebook.add(tab1, text='Стройка')

tab2 = ttk.Frame(notebook)
notebook.add(tab2, text='Фарм')

tab3 = ttk.Frame(notebook)
notebook.add(tab3, text='Вкладка 3')

"""вкладка "Фарм"""
# строка 1 - Отправка группы войск в новом окне
troops_1_4_5_icon = ImageTk.PhotoImage(Image.open("images/1+4+5.png"))
btn_send_troops_1_4_5 = tk.Button(tab2, image=troops_1_4_5_icon , command=lambda: send_troops_new_window('1', '4', '5'))
btn_send_troops_1_4_5.grid(row=1, column=1)

# строка 2 - Отправка одного вида войск в новом окне
if Variables.race == 'Romans':
    troops1_icon = ImageTk.PhotoImage(Image.open("images/Legioner_icon.png"))
else:
    troops1_icon = ImageTk.PhotoImage(Image.open("images/Mercenaries_ico.png"))
btn_send_troops1 = tk.Button(tab2, image=troops1_icon, command=lambda: send_troops_new_window('1'))
btn_send_troops1.grid(row=2, column=1)

if Variables.race == 'Romans':
    troops2_icon = ImageTk.PhotoImage(Image.open("images/Praetorian_icon.png"))
else:
    troops2_icon = ImageTk.PhotoImage(Image.open("images/Bowman_icon.png"))
btn_send_troops2 = tk.Button(tab2, image=troops2_icon, command=lambda: send_troops_new_window('2'))
btn_send_troops2.grid(row=2, column=2)

if Variables.race == 'Romans':
    troops3_icon = ImageTk.PhotoImage(Image.open("images/Imperian_icon.png"))
else:
    troops3_icon = ImageTk.PhotoImage(Image.open("images/Spotter_icon.png"))
btn_send_troops3 = tk.Button(tab2, image=troops3_icon, command=lambda: send_troops_new_window('3'))
btn_send_troops3.grid(row=2, column=3)

if Variables.race == 'Romans':
    troops4_icon = ImageTk.PhotoImage(Image.open("images/Legati_icon.png"))
else:
    troops4_icon = ImageTk.PhotoImage(Image.open("images/Steppe.png"))
btn_send_troops4 = tk.Button(tab2, image=troops4_icon, command=lambda: send_troops_new_window('4'))
btn_send_troops4.grid(row=2, column=4)

if Variables.race == 'Romans':
    troops5_icon = ImageTk.PhotoImage(Image.open("images/Imperatoris_icon.png"))
else:
    troops5_icon = ImageTk.PhotoImage(Image.open("images/Marksman_ico.png"))
btn_send_troops5 = tk.Button(tab2, image=troops5_icon, command=lambda: send_troops_new_window('5'))
btn_send_troops5.grid(row=2, column=5)

if Variables.race == 'Romans':
    troops6_icon = ImageTk.PhotoImage(Image.open("images/Caesaris_icon.png"))
else:
    troops6_icon = ImageTk.PhotoImage(Image.open("images/Marauder_ico.png"))
btn_send_troops6 = tk.Button(tab2, image=troops6_icon, command=lambda: send_troops_new_window('6'))
btn_send_troops6.grid(row=2, column=6)

troops_hero_icon = ImageTk.PhotoImage(Image.open("images/Hero_icon.png"))
btn_send_troops_hero = tk.Button(tab2, image=troops_hero_icon, command=lambda: send_troops_new_window('11'))
btn_send_troops_hero.grid(row=2, column=7)

# строка 3 - Количество воск для отправки или добавления в фармлист
ent_troops1 = tk.Entry(tab2, width=3)
ent_troops1.grid(row=3, column=1)
ent_troops1.insert(0, "2")

ent_troops2 = tk.Entry(tab2, width=3)
ent_troops2.grid(row=3, column=2)
ent_troops2.insert(0, "2")

ent_troops3 = tk.Entry(tab2, width=3)
ent_troops3.grid(row=3, column=3)
ent_troops3.insert(0, "2")

ent_troops4 = tk.Entry(tab2, width=3)
ent_troops4.grid(row=3, column=4)
ent_troops4.insert(0, "1")

troops5 = tk.Entry(tab2, width=3)
troops5.grid(row=3, column=5)
troops5.insert(0, "1")

troops6 = tk.Entry(tab2, width=3)
troops6.grid(row=3, column=6)
troops6.insert(0, "1")

# строка 4 - Статус добавления в фармлист
lable_add_to_farmlist_status = tk.Label(tab2, text="Добавить в фармлист", width=31, anchor='w')
lable_add_to_farmlist_status.grid(row=4, column=1, columnspan=10)

# строка 5 - Кнопки добавления одного вида войск в фармлист
add_to_farmlist_icon = ImageTk.PhotoImage(Image.open("images/Add_to_farmlist_icon.png"))

btn_add_to_farmlist_troops1 = tk.Button(tab2, image=add_to_farmlist_icon, command=lambda: add_to_farmlist('1'))
btn_add_to_farmlist_troops1.grid(row=5, column=1)

btn_add_to_farmlist_troops2 = tk.Button(tab2, image=add_to_farmlist_icon, command=lambda: add_to_farmlist('2'))
btn_add_to_farmlist_troops2.grid(row=5, column=2)

btn_add_to_farmlist_troops3 = tk.Button(tab2, image=add_to_farmlist_icon, command=lambda: add_to_farmlist('3'))
btn_add_to_farmlist_troops3.grid(row=5, column=3)

btn_add_to_farmlist_troops4 = tk.Button(tab2, image=add_to_farmlist_icon, command=lambda: add_to_farmlist('4'))
btn_add_to_farmlist_troops4.grid(row=5, column=4)

btn_add_to_farmlist_troops5 = tk.Button(tab2, image=add_to_farmlist_icon, command=lambda: add_to_farmlist('5'))
btn_add_to_farmlist_troops5.grid(row=5, column=5)

btn_add_to_farmlist_troops6 = tk.Button(tab2, image=add_to_farmlist_icon, command=lambda: add_to_farmlist('6'))
btn_add_to_farmlist_troops6.grid(row=5, column=6)

# строка 6 - Кнопки добавления группы войск в фармлист
add_to_farm_troops_1_4_5_icon = ImageTk.PhotoImage(Image.open("images/farm_1+4+5.png"))
btn_add_to_farm_troops_1_4_5_icon = tk.Button(tab2, image=add_to_farm_troops_1_4_5_icon , command=lambda: add_to_farmlist('1', '4', '5'))
btn_add_to_farm_troops_1_4_5_icon.grid(row=6, column=1)

# строка 7 - Статус добавления в фармлист
lable_send_after_time = tk.Label(tab2, text="Отправить через ... секунд", width=31, anchor='w')
lable_send_after_time.grid(row=7, column=1, columnspan=10)

# строка 7 - Фармлист статус
lable_message = tk.Label(tab2, text="...", width=31, anchor='w')
lable_message.grid(row=7, column=1, columnspan=10)


# строка 8 - Фармлист
farmlist_icon = ImageTk.PhotoImage(Image.open("images/farmlist_icon.png"))
btn_farmlist = tk.Button(tab2, image=farmlist_icon, command=lambda: start_farmlist('Farmlist'))
btn_farmlist.grid(row=8, column=1)

btn_farm_wait_times_to_repeat = tk.Entry(tab2, width=3)
btn_farm_wait_times_to_repeat.grid(row=8, column=2)
btn_farm_wait_times_to_repeat.insert(0, "3")
lable_btn_farm_wait_times_to_repeat = tk.Label(tab2, text="repeat", width=7, anchor='w')
lable_btn_farm_wait_times_to_repeat.grid(row=8, column=3, columnspan=1)

go_to_next_farm = tk.BooleanVar()
go_to_next_farm.set(True)
check_go_to_next_farm = tk.Checkbutton(tab2, variable=go_to_next_farm, offvalue=False, onvalue=True)
check_go_to_next_farm.grid(row=8, column=4)
lable_check_go_to_next_farm = tk.Label(tab2, text="next farm", width=7, anchor='w')
lable_check_go_to_next_farm.grid(row=8, column=5, columnspan=1)

ent_farm_circles = tk.Entry(tab2, width=3)
ent_farm_circles.grid(row=8, column=6)
ent_farm_circles.insert(0, "1")
lable_ent_farm_circles = tk.Label(tab2, text="circles", width=7, anchor='w')
lable_ent_farm_circles.grid(row=8, column=7, columnspan=1)

# строка 9 - Оазисы статус
lable_oasises = tk.Label(tab2, text="Статус по оазисам", width=31, anchor='w')
lable_oasises.grid(row=9, column=1, columnspan=10)

# строка 10 - Оазисы
oasises_refresh_icon = ImageTk.PhotoImage(Image.open("images/Oasises_refresh.png"))
btn_oasises_refresh = tk.Button(tab2, image=oasises_refresh_icon, command=lambda: (Oasises.refresh_oasises(driver, active_village),update_label_message(lable_oasises, "Оазисы обновлены")))
btn_oasises_refresh.grid(row=10, column=1)


farm_oasises_icon = ImageTk.PhotoImage(Image.open("images/Oasises_start_farm.png"))
btn_farm_oasises = tk.Button(tab2, image=farm_oasises_icon, command=lambda: (start_farmlist('FarmOasises'), update_label_message(lable_oasises, "Фарм на оазисы отправлен")))
btn_farm_oasises.grid(row=10, column=2)

scan_oasises_icon = ImageTk.PhotoImage(Image.open("images/Scan_oasises.png"))
btn_scan_oasises = tk.Button(tab2, image=scan_oasises_icon, command=lambda: (Oasises.scan_oasises(driver), update_label_message(lable_oasises, "Оазисы отсканированы")))
btn_scan_oasises.grid(row=10, column=7)

# строка 20 - Кнопки отправки войск через время
troops_watch_icon = ImageTk.PhotoImage(Image.open("images/Watch_icon.png"))
btn_send_troops1_watch = tk.Button(tab2, image=troops_watch_icon, command=lambda: btn_send_troops_timer('1'))
btn_send_troops1_watch.grid(row=20, column=1)

# строка 21 - Секунды для отправки войск через время
troops1_sender_timer = tk.Entry(tab2, width=3)
troops1_sender_timer.grid(row=21, column=1)
troops1_sender_timer.insert(0, "0")

troops5_sender_timer = tk.Entry(tab2, width=3)
troops5_sender_timer.grid(row=21, column=5)
troops5_sender_timer.insert(0, "0")

troops_hero_sender_timer = tk.Entry(tab2, width=3)
troops_hero_sender_timer.grid(row=21, column=7)
troops_hero_sender_timer.insert(0, "0")

"""Стройка"""
# строка 1 - Запуск браузера, обновить отчёты
open_browser_button_icon = ImageTk.PhotoImage(Image.open("images/Chrome_icon.png"))
open_browser_button = tk.Button(tab1, image=open_browser_button_icon, command=open_browser)
open_browser_button.grid(row=1, column=1)

reports_icon = ImageTk.PhotoImage(Image.open("images/reports_icon.png"))
btn_refresh_reports_excel = tk.Button(tab1, image=reports_icon, command=lambda: (on_button_refresh_reports_excel_click(), update_label_message(lable_reports, "Отчёты обновлены")))
btn_refresh_reports_excel.grid(row=1, column=2)

num_of_pages_refresh_reports_excel = tk.Entry(tab1, width=3)
num_of_pages_refresh_reports_excel.grid(row=1, column=3)
num_of_pages_refresh_reports_excel.insert(0, "50")

lable_reports = tk.Label(tab1, text="Статус по отчётам", width=15, anchor='w')
lable_reports.grid(row=1, column=4, columnspan=3)

# строительство
btn_build_after_or_in_time_icon = ImageTk.PhotoImage(Image.open("images/Build_after_time_icon.png"))
btn_build_after_or_in_time = tk.Button(tab1, image=btn_build_after_or_in_time_icon, command=on_btn_build_after_or_in_time)
btn_build_after_or_in_time.grid(row=3, column=1)

Entry_build_after_time = tk.Entry(tab1, width=3)
Entry_build_after_time.grid(row=3, column=2)
Entry_build_after_time.insert(0, "1")

Entry_build_in_time = tk.Entry(tab1, width=5)
Entry_build_in_time.grid(row=3, column=3)
Entry_build_in_time.insert(0, "00:00")

building_queue_listbox = tk.Listbox(tab1, height=10, width=30)
building_queue_listbox.grid(row=4, column=1, columnspan=7)

win.mainloop()