import math
import random
import re
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup

import Variables


def parsing_rally_point(driver):
    # парсинг атак
    list_rally_point_attack = []
    driver.get(Variables.server+'/build.php?gid=16&tt=1&filter=2&subfilters=4')
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    num_of_attack = soup.find('h4', class_="spacer").get_text()
    num_of_attack = re.search(r'\((\d+)\)', num_of_attack).group(1)
    attack_pages = math.ceil(int(num_of_attack) / 10)

    for page in range(1, attack_pages + 1):
        if page != 1:
            driver.get(
                Variables.server+'/build.php?gid=16&tt=1&filter=2&subfilters=4&sortOrder=earliestArrivingFirst&page=' + str(
                    page))
            html_content = driver.page_source
            soup = BeautifulSoup(html_content, 'html.parser')

        items = soup.find_all('table', class_=['troop_details outRaid', 'troop_details outAttack'])

        for i in items:
            map_id = i.find('td', class_='troopHeadline').find('a').get('href')
            map_id = map_id.split('d=')[1]
            at = i.find('div', class_='at').find('span').get_text()
            at = re.search(r'\d{2}:\d{2}:\d{2}', at).group(0)
            list_rally_point_attack.append(map_id)
            list_rally_point_attack.append(at)
        time.sleep(random.randint(1, 2))

    # парсинг возвращений
    list_rally_point_return = []
    driver.get(Variables.server+'/build.php?gid=16&tt=1&filter=1&subfilters=2,3')
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    num_of_attack = soup.find('h4', class_="spacer").get_text()
    num_of_attack = re.search(r'\((\d+)\)', num_of_attack).group(1)
    attack_pages = math.ceil(int(num_of_attack) / 10)

    for page in range(1, attack_pages + 1):
        if page != 1:
            driver.get(
                Variables.server+'/build.php?gid=16&tt=1&filter=1&subfilters=2%2C3&page=' + str(page))
            html_content = driver.page_source
            soup = BeautifulSoup(html_content, 'html.parser')

        items = soup.find_all('table', class_='troop_details inReturn')

        for i in items:
            map_id = i.find('td', class_='troopHeadline').find('a').get('href')
            map_id = map_id.split('d=')[1]
            at = i.find('div', class_='at').find('span').get_text()
            at = re.search(r'\d{2}:\d{2}:\d{2}', at).group(0)
            list_rally_point_return.append(map_id)
            list_rally_point_return.append(at)
        time.sleep(random.randint(1, 2))

    # записать атаки и возвращения в Уксель
    workbook = load_workbook(filename='BadrBot.xlsx')
    sheet = workbook['Rally Point']

    for row in range(1, 501):  # очищаем 500 строк
        for col in range(1, 5):
            cell = sheet.cell(row=row, column=col)
            cell.value = None

    col = 1
    row = 1
    for attack in list_rally_point_attack:
        sheet.cell(row=row, column=col).value = attack
        if col < 2:
            col += 1
        else:
            col = 1
            row += 1
    col = 3
    row = 1
    for ret in list_rally_point_return:
        sheet.cell(row=row, column=col).value = ret
        if col < 4:
            col += 1
        else:
            col = 3
            row += 1

    workbook.save(filename='BadrBot.xlsx')