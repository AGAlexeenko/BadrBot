from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import Variables

def scan_oasises(driver):
    wait = WebDriverWait(driver, 15, poll_frequency=1)
    radius_cells = 3
    x_start = 92
    y_start = -177
    confirm_locator = ("xpath", '//a[@class="a arrow"]')
    data = {}
    workbook = load_workbook(filename='BadrBot.xlsx')
    sheet = workbook['Oasises']
    for row in range(4, 100):  # очищаем 100 строк
        for col in range(1, 25):
            cell = sheet.cell(row=row, column=col)
            cell.value = None

    for x in range(x_start-radius_cells, x_start+radius_cells+1):
        for y in range(y_start*-1-radius_cells, (y_start*-1+radius_cells+1)):
            y = y*-1
            driver.get(f'{Variables.server}/karte.php?x={x}&y={y}')
            wait.until(EC.visibility_of_element_located(confirm_locator))
            title = driver.find_element("xpath", '//h1[@class ="titleInHeader"]').text
            title = title.split('(')[0]
            if title == 'Unoccupied oasis ‭':
                try:
                    html_content = driver.page_source
                    soup = BeautifulSoup(html_content, 'html.parser')
                    links_with_mapid = soup.find('a', href=lambda x: x and 'targetMapId' in x)
                    href_value_get = links_with_mapid.get('href')
                    map_id = href_value_get.split('targetMapId=')[1]
                    table = soup.find('table', id='troop_info')
                    rows = table.find_all('tr')
                    distance = round(((x - x_start) ** 2 + (y - y_start) ** 2))
                    data = {"coordinates": '0|0', 'x': 0, 'y': 0, 'distance': 0, 'Rats':0, 'Spiders':0, 'Snakes':0, 'Bats':0, 'Wild Boars':0, 'Wolves':0, 'Bears':0, 'Crocodiles':0, 'Tigers':0, 'Elephants':0}
                    data['coordinates'] = str(x)+'|'+str(y)
                    data['x'] = x
                    data['y'] = y
                    data['distance'] = distance
                    data['title'] = title
                    data['map_id'] = map_id
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) == 3:
                            description = cells[2].get_text(strip=True)
                            value = cells[1].get_text(strip=True)
                            data[description] = value

                    last_row = sheet.max_row
                    for row in range(sheet.max_row, 1, -1):
                        if sheet.cell(row=row, column=3).value is not None:
                            last_row = row
                            break
                    sheet['A' + str(last_row + 1)] = data['map_id']
                    sheet['B' + str(last_row + 1)] = data['title']
                    sheet['C' + str(last_row + 1)] = data['coordinates']
                    sheet['D' + str(last_row + 1)] = data['x']
                    sheet['E' + str(last_row + 1)] = data['y']
                    sheet['N' + str(last_row + 1)] = data['distance']
                    sheet['O' + str(last_row + 1)] = data['Rats']
                    sheet['P' + str(last_row + 1)] = data['Spiders']
                    sheet['Q' + str(last_row + 1)] = data['Snakes']
                    sheet['R' + str(last_row + 1)] = data['Bats']
                    sheet['S' + str(last_row + 1)] = data['Wild Boars']
                    sheet['T' + str(last_row + 1)] = data['Wolves']
                    sheet['U' + str(last_row + 1)] = data['Bears']
                    sheet['V' + str(last_row + 1)] = data['Crocodiles']
                    sheet['W' + str(last_row + 1)] = data['Tigers']
                    sheet['X' + str(last_row + 1)] = data['Elephants']
                    driver.find_element("xpath", '//div[@class="dialogCancelButton iconButton small cancel"]').click()
                    workbook.save(filename='BadrBot.xlsx')
                except NoSuchElementException:
                    continue
            else:
                continue

def refresh_oasises(driver, active_village):
    wait = WebDriverWait(driver, 15, poll_frequency=1)
    workbook = load_workbook(filename='BadrBot.xlsx')
    sheet = workbook['FarmOasises']
    list_of_cordinates_xl = []

    last_row = sheet.max_row
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row=row, column=4).value is not None:
            last_row = row
            break

    for row in range(4, last_row + 1):  # считать все данные с Эксель листа
        x_xl = sheet.cell(row=row, column=5).value
        y_xl = sheet.cell(row=row, column=6).value
        list_of_cordinates_xl.append([x_xl, y_xl])

    confirm_locator = ("xpath", '//a[@class="a arrow"]')
    data = {}

    sheet = workbook['Oasises']
    for row in range(4, 100):  # очищаем 100 строк
        for col in range(1, 25):
            cell = sheet.cell(row=row, column=col)
            cell.value = None

    x_home = driver.find_element("xpath", f'//span[@class="coordinatesGrid" and @data-did="{str(active_village[0])}"]').get_attribute(
        'data-x')
    y_home = driver.find_element("xpath", f'//span[@class="coordinatesGrid" and @data-did="{str(active_village[0])}"]').get_attribute(
        'data-y')

    for xy in list_of_cordinates_xl:
        x = int(xy[0])
        y = int(xy[1])
        driver.get(f'{Variables.server}/karte.php?x={x}&y={y}')
        wait.until(EC.visibility_of_element_located(confirm_locator))
        title = driver.find_element("xpath", '//h1[@class ="titleInHeader"]').text
        title = title.split('(')[0]
        if title == 'Unoccupied oasis ‭':
            try:
                html_content = driver.page_source
                soup = BeautifulSoup(html_content, 'html.parser')
                links_with_mapid = soup.find('a', href=lambda x: x and 'targetMapId' in x)
                href_value_get = links_with_mapid.get('href')
                map_id = href_value_get.split('targetMapId=')[1]
                table = soup.find('table', id='troop_info')
                rows = table.find_all('tr')
                distance = round(((x - int(x_home)) ** 2 + (y - int(y_home)) ** 2))
                data = {"coordinates": '0|0', 'x': 0, 'y': 0, 'distance': 0, 'Rats':0, 'Spiders':0, 'Snakes':0, 'Bats':0, 'Wild Boars':0, 'Wolves':0, 'Bears':0, 'Crocodiles':0, 'Tigers':0, 'Elephants':0}
                data['coordinates'] = str(x)+'|'+str(y)
                data['x'] = x
                data['y'] = y
                data['distance'] = distance
                data['title'] = title
                data['map_id'] = map_id
                for row in rows:
                    cells = row.find_all('td')
                    if len(cells) == 3:
                        description = cells[2].get_text(strip=True)
                        value = cells[1].get_text(strip=True)
                        data[description] = value

                last_row = sheet.max_row
                for row in range(sheet.max_row, 1, -1):
                    if sheet.cell(row=row, column=3).value is not None:
                        last_row = row
                        break

                sheet['A' + str(last_row + 1)] = data['map_id']
                sheet['B' + str(last_row + 1)] = data['title']
                sheet['C' + str(last_row + 1)] = data['coordinates']
                sheet['D' + str(last_row + 1)] = data['x']
                sheet['E' + str(last_row + 1)] = data['y']
                sheet['N' + str(last_row + 1)] = data['distance']
                sheet['O' + str(last_row + 1)] = data['Rats']
                sheet['P' + str(last_row + 1)] = data['Spiders']
                sheet['Q' + str(last_row + 1)] = data['Snakes']
                sheet['R' + str(last_row + 1)] = data['Bats']
                sheet['S' + str(last_row + 1)] = data['Wild Boars']
                sheet['T' + str(last_row + 1)] = data['Wolves']
                sheet['U' + str(last_row + 1)] = data['Bears']
                sheet['V' + str(last_row + 1)] = data['Crocodiles']
                sheet['W' + str(last_row + 1)] = data['Tigers']
                sheet['X' + str(last_row + 1)] = data['Elephants']
                driver.find_element("xpath", '//div[@class="dialogCancelButton iconButton small cancel"]').click()
                workbook.save(filename='BadrBot.xlsx')
            except NoSuchElementException:
                continue
        else:
            continue
